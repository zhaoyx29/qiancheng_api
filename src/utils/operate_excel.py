# -*- coding: utf-8 -*-
# @Author  :  'zyx'
# @Email   : 458757014@qq.com
# @File    : operate_excel.py
# @Software: PyCharm Community Edition

from openpyxl import load_workbook

class OperateExcel:
    def __init__(self,filename,sheetname):
        #打开Excel文件，定位sheet
        self.filename = filename
        self.wb = load_workbook(filename)
        self.sh = self.wb[sheetname]

    #读取Excel中的数据
    def read_data(self):
        read_datas=[]    #存储读取的所有测试数据
        #读取Excel中的数据，每一组数据组成一个字典，作为测试数据
        for i in range(2,self.sh.max_row+1):
            dict_data={}
            dict_data['case_id']=self.sh.cell(i,column=1).value
            dict_data['func']=self.sh.cell(i,column=2).value
            dict_data['describe']=self.sh.cell(i,column=3).value
            dict_data['method']=self.sh.cell(i,column=4).value
            dict_data['apiName']=self.sh.cell(i,column=5).value
            dict_data['param']=self.sh.cell(i,column=6).value
            dict_data['expected']=self.sh.cell(i,column=7).value
            read_datas.append(dict_data)
        return read_datas

    def write_back(self,row,column,new_value):
        self.sh.cell(row,column).value=new_value
        self.wb.save(self.filename)

from src.test_resource.common.filepath import *
if __name__ == '__main__':
    data=OperateExcel(data_path+r'\test_datas.xlsx','test_data').read_data()
    print(type(data[0]))
    print(data[0])
